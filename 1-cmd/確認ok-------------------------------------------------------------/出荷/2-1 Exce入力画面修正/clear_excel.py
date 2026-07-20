from pathlib import Path
import openpyxl

# =========================
# Excel文件夹
# =========================

excel_folder = Path(
    r"C:\c_wk\10_会社\PDF-相关\Test\Test19"
)

sheet_name = "入力"

keyword = "費"

success = 0
skip = 0


# =========================
# 批量处理
# =========================

for excel_file in excel_folder.glob("*.xlsx"):

    if excel_file.name.startswith("~$"):
        continue

    try:

        print()
        print("处理:", excel_file.name)

        wb = openpyxl.load_workbook(excel_file)

        if sheet_name not in wb.sheetnames:

            print(
                "找不到Sheet:",
                sheet_name
            )

            skip += 1
            wb.close()
            continue

        ws = wb[sheet_name]

        found_row = None

        # =========================
        # 查找 B13:B24
        # B:E 为合并单元格
        # =========================

        for row in range(13, 25):

            row_text = []

            for col in range(1, 8):

                value = ws.cell(row, col).value

                if value is not None:

                    print(
                        row,
                        col,
                        repr(str(value))
                    )

                    if "費" in str(value):

                        found_row = row
                        print("找到費:", row, col)

                        break

            if found_row:
                break

        # =========================
        # 清空 A:G
        # found_row ～ 24行
        # =========================

        if found_row:

            print(
                f"找到費: 第 {found_row} 行"
            )

            merged_ranges = list(
                ws.merged_cells.ranges
            )

            for row in range(
                found_row,
                25
            ):

                for col in range(
                    1,
                    8
                ):   # A:G

                    cell = ws.cell(
                        row,
                        col
                    )

                    skip_cell = False

                    for mr in merged_ranges:

                        if cell.coordinate in mr:

                            if (
                                cell.coordinate
                                != mr.start_cell.coordinate
                            ):

                                skip_cell = True
                                break

                    if not skip_cell:

                        cell.value = None

            print(
                f"已清空 A{found_row}:G24"
            )

        else:

            print(
                "B13:B24 未找到 費"
            )


        wb.save(excel_file)

        wb.close()

        success += 1

    except Exception as e:

        print(
            "错误:",
            excel_file.name,
            e
        )

        skip += 1


print()
print("===================")
print("完成:", success)
print("跳过:", skip)
print("===================")

input("按 Enter 结束...")