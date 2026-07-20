from pathlib import Path
import openpyxl

# =========================
# Excel文件夹
# =========================

excel_folder = Path(
    r"C:\c_wk\10_会社\PDF-相关\Test\Test19"
)

success = 0
skip = 0


for excel_file in excel_folder.glob("*.xlsx"):

    if excel_file.name.startswith("~$"):
        continue

    try:

        wb = openpyxl.load_workbook(excel_file)

        if "明細" not in wb.sheetnames:

            print(
                "找不到Sheet[明細]:",
                excel_file.name
            )

            skip += 1
            wb.close()
            continue


        ws = wb["明細"]


        # =========================
        # 1. 删除H列以后所有列
        # =========================

        if ws.max_column > 7:

            ws.delete_cols(
                8,
                ws.max_column - 7
            )


        # =========================
        # 2. 删除D列空白以下所有行
        # =========================

        for row in range(
            7,
            ws.max_row + 1
        ):

            value = ws.cell(
                row,
                4
            ).value

            if (
                value is None
                or str(value).strip() == ""
            ):

                ws.delete_rows(
                    row,
                    ws.max_row - row + 1
                )

                break


        # =========================
        # 3. B2:I2
        # 株式会社ABC → ABC
        # =========================

        company_name = ws["B2"].value

        if company_name:

            ws["B2"] = (
                str(company_name)
                .replace("御見積明細書", "部品明細書")
                .strip()
            )

        # =========================
        # 4. 设置列宽
        # =========================

        ws.column_dimensions["D"].width = 30.64
        ws.column_dimensions["E"].width = 30.64

        ws.column_dimensions["F"].width = 7.64
        ws.column_dimensions["G"].width = 7.64

        wb.save(excel_file)

        wb.close()

        success += 1

        print(
            "完成:",
            excel_file.name
        )


    except Exception as e:

        skip += 1

        print(
            "错误:",
            excel_file.name,
            e
        )


print()
print("===================")
print("完成:", success)
print("跳过:", skip)
print("===================")

input("按 Enter 结束...")