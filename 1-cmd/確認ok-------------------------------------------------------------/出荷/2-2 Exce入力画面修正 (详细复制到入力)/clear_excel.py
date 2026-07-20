from pathlib import Path
import openpyxl


# =========================
# Excel文件夹
# =========================

excel_folder = Path(
    r"C:\c_wk\10_会社\PDF-相关\Test\Test19"
)


success = 0
skip = 0


# =========================
# 开始处理
# =========================

for excel_file in excel_folder.glob("*.xlsx"):

    # 跳过临时文件
    if excel_file.name.startswith("~$"):
        continue

    try:

        wb = openpyxl.load_workbook(
            excel_file
        )


        # Sheet名称
        if "明細" not in wb.sheetnames:

            print(
                "找不到Sheet[明細]:",
                excel_file.name
            )

            skip += 1
            wb.close()
            continue


        if "入力" not in wb.sheetnames:

            print(
                "找不到Sheet[入力]:",
                excel_file.name
            )

            skip += 1
            wb.close()
            continue


        ws_meisai = wb["明細"]
        ws_input = wb["入力"]


        # =========================
        # B8:G19 → 入力
        # =========================

        for i in range(12):

            src_row = 8 + i
            dst_row = 13 + i


            # D列为空则停止
            check_value = ws_meisai.cell(
                src_row,
                4
            ).value


            if (
                check_value is None
                or str(check_value).strip() == ""
            ):

                # 清空剩余区域
                for clear_row in range(
                    dst_row,
                    25
                ):

                    ws_input.cell(
                        clear_row,
                        1
                    ).value = None

                    ws_input.cell(
                        clear_row,
                        2
                    ).value = None

                    ws_input.cell(
                        clear_row,
                        6
                    ).value = None

                    ws_input.cell(
                        clear_row,
                        7
                    ).value = None

                break


            # =========================
            # B → A
            # =========================

            ws_input.cell(
                dst_row,
                1
            ).value = ws_meisai.cell(
                src_row,
                2
            ).value


            # =========================
            # D:E → B:E
            # =========================

            ws_input.cell(
                dst_row,
                2
            ).value = ws_meisai.cell(
                src_row,
                4
            ).value


            # =========================
            # F → F
            # =========================

            ws_input.cell(
                dst_row,
                6
            ).value = ws_meisai.cell(
                src_row,
                6
            ).value


            # =========================
            # G → G
            # =========================

            ws_input.cell(
                dst_row,
                7
            ).value = ws_meisai.cell(
                src_row,
                7
            ).value


        wb.save(excel_file)

        wb.close()

        success += 1

        print(
            "完成:",
            excel_file.name
        )


    except Exception as e:

        skip += 1

        print(
            "错误:",
            excel_file.name,
            e
        )


print()
print("===================")
print("完成:", success)
print("跳过:", skip)
print("===================")

input("按 Enter 结束...")